import asyncio
import urllib.parse
import openpyxl
import os
import re
import random
import time
from playwright.async_api import async_playwright, Error as PlaywrightError

MAX_FIRM_RETRIES = 2
MAX_PAGE_RETRIES = 3
SAVE_EVERY_N = 5
CONCURRENCY_LIMIT = 4
GRID_STEPS = 7
ZOOM = 13
LAT_MIN, LAT_MAX = 55.1, 56.2
LON_MIN, LON_MAX = 36.7, 38.4

CAPTCHA_MARKERS = [
    "подтвердите, что вы не робот", "капча", "captcha", "unusual traffic", "доступ ограничен",
    "just a moment", "checking your browser", "access denied", "attention required",
    "please verify you are a human", "too many requests",
]
CAPTCHA_STATUS_CODES = {403, 429}


def get_firm_id(url):
    if not url:
        return None
    match = re.search(r'/firm/(\d+)', str(url))
    return match.group(1) if match else None


def generate_grid(lat_min, lat_max, lon_min, lon_max, steps):
    lat_step = (lat_max - lat_min) / steps
    lon_step = (lon_max - lon_min) / steps
    points = []
    for i in range(steps):
        for j in range(steps):
            lat = round(lat_min + lat_step * (i + 0.5), 6)
            lon = round(lon_min + lon_step * (j + 0.5), 6)
            col_letter = chr(65 + j)
            row_num = steps - i
            points.append((lat, lon, f"Сектор {col_letter}{row_num}"))
    points.sort(key=lambda x: x[2])
    return points


def normalize_text(t):
    return re.sub(r'\s+', ' ', (t or '')).strip().lower()


async def get_rubric_text(page):
    try:
        rubric = await page.evaluate("""() => {
            const candidates = Array.from(document.querySelectorAll('[class*="rubric" i], [class*="Rubric" i], a[href*="/rubric/"]'));
            for (const el of candidates) {
                const txt = (el.innerText || '').trim();
                if (txt && txt.length < 120) return txt;
            }
            const main = document.querySelector('main') || document.body;
            return (main.innerText || '').slice(0, 600);
        }""")
        return normalize_text(rubric)
    except Exception:
        return ""


def matches_keywords(rubric_text, title_text, keywords):
    if not keywords:
        return True
    haystack = f"{rubric_text} {title_text}".lower()
    return any(kw.lower().strip() in haystack for kw in keywords if kw.strip())


async def page_looks_like_captcha(page):
    try:
        text = normalize_text(await page.evaluate("() => document.body.innerText.slice(0, 2000)"))
    except Exception:
        return False
    return any(marker in text for marker in CAPTCHA_MARKERS)


class Job:
    """In-memory state for one parser run - one instance per niche card click."""

    def __init__(self, job_id, category, description, queries, out_dir):
        self.id = job_id
        self.category = category
        self.description = description
        self.queries = queries  # [{"query": str, "keywords": [str]}]
        self.out_dir = out_dir
        self.status = "queued"  # queued|running|captcha|done|error
        self.log_lines = []
        self.stats = {"collected": 0, "duplicates": 0, "filtered_out": 0, "errors": 0}
        self.raw_path = os.path.join(out_dir, "raw.xlsx")
        self.dedup_path = os.path.join(out_dir, "dedup.xlsx")
        self.archive_path = os.path.join(out_dir, "archive.zip")
        self.on_captcha = None  # async callback(job) -> None, set by the API layer
        self.task = None  # asyncio.Task running this job, set by worker_loop - used to cancel
        self.cancelled = False  # set when cancelled while still queued (before a task exists)

    def log(self, msg):
        line = f"[{time.strftime('%H:%M:%S')}] {msg}"
        self.log_lines.append(line)
        print(f"[{self.id}] {line}", flush=True)


async def process_firm(context, url, ws, wb, lock, semaphore, job, keywords):
    async with semaphore:
        for attempt in range(MAX_FIRM_RETRIES + 1):
            page = None
            try:
                page = await context.new_page()
                await page.route("**/*", lambda route: route.abort()
                    if route.request.resource_type in ["image", "stylesheet", "media", "font"]
                    else route.continue_())
                await page.goto(url, wait_until="domcontentloaded", timeout=25000)

                h1 = page.locator('h1')
                title = (await h1.first.inner_text()).split('\n')[0].strip() if await h1.count() > 0 else "Без названия"

                rubric_text = await get_rubric_text(page)
                if not matches_keywords(rubric_text, title, keywords):
                    await page.close()
                    async with lock:
                        job.stats["filtered_out"] += 1
                    return False

                geo = page.locator('a[href*="/geo/"]')
                address = (await geo.first.inner_text()).replace('\n', ', ').strip() if await geo.count() > 0 else "Не указан"

                await page.evaluate("""() => {
                    const btns = Array.from(document.querySelectorAll('button, div, span'));
                    btns.forEach(b => {
                        const t = (b.innerText || '').toLowerCase();
                        if (t.includes('показать телефон') || t.includes('соцсети') || t.includes('показать контакт')) {
                            try { b.click(); } catch(e) {}
                        }
                    });
                }""")
                await page.wait_for_timeout(random.randint(500, 1000))

                tels = page.locator('a[href^="tel:"]')
                phones = []
                for i in range(await tels.count()):
                    href = await tels.nth(i).get_attribute('href')
                    if href:
                        num = href.replace('tel:', '').strip()
                        if num and num not in phones:
                            phones.append(num)
                phone_str = ", ".join(phones) if phones else "Не указан"

                raw_data = await page.evaluate("""() => {
                    let results = [];
                    document.querySelectorAll('a').forEach(a => { if (a.href) results.push(a.href); });
                    return Array.from(new Set(results));
                }""")

                bad_domains = ["2gis", "yandex", "google", "flamp", "apple", "play.google", "otello", "restoclub", "tomesto", "zoon", "sbermarket", "megamarket", "delivery-club", "eda.yandex", "mos.ru", "gosuslugi", "w3.org", "github.com", "yoo.money"]
                good_indicators = [".ru", ".com", ".рф", ".org", ".net", ".info", ".su", ".pro", ".moscow", ".agency", ".media", ".digital", "vk.com", "t.me", "wa.me", "taplink.cc"]
                websites = []
                for val in raw_data:
                    val_lower = val.lower()
                    if "redirect?url=" in val_lower:
                        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(val).query)
                        if "url" in parsed:
                            val = urllib.parse.unquote(parsed["url"][0])
                            val_lower = val.lower()
                    if any(ind in val_lower for ind in good_indicators) and not any(bd in val_lower for bd in bad_domains):
                        if "@" not in val and val not in websites:
                            websites.append(val if val_lower.startswith("http") else "https://" + val)

                real_sites = [w for w in websites if not any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
                socials = [w for w in websites if any(s in w.lower() for s in ['vk.com', 't.me', 'instagram.com', 'wa.me', 'whatsapp.com'])]
                site_str = real_sites[0] if real_sites else (socials[0] if socials else "Нет сайта")

                site_label = "[WEB]"
                if "t.me" in site_str:
                    site_label = "[TG]"
                elif "vk.com" in site_str:
                    site_label = "[VK]"
                elif "wa.me" in site_str or "whatsapp" in site_str:
                    site_label = "[WA]"
                elif site_str == "Нет сайта":
                    site_label = "[---]"

                async with lock:
                    job.stats["collected"] += 1
                    curr_no = job.stats["collected"]
                    ws.append([curr_no, title, address, phone_str, site_str, site_label, url])
                    if curr_no % SAVE_EVERY_N == 0:
                        wb.save(job.raw_path)
                    job.log(f"[{curr_no}] {title} | {phone_str} | {site_label} {site_str}")

                return True

            except asyncio.CancelledError:
                raise
            except (PlaywrightError, Exception) as e:
                if attempt < MAX_FIRM_RETRIES:
                    await asyncio.sleep(1.5 * (attempt + 1))
                    continue
                async with lock:
                    job.stats["errors"] += 1
                return False
            finally:
                if page:
                    try:
                        await page.close()
                    except Exception:
                        pass
        return False


async def collect_valid_hrefs(page):
    cards = page.locator('a[href*="/firm/"]')
    count = await cards.count()
    hrefs = []
    for i in range(count):
        href = await cards.nth(i).get_attribute('href')
        if href:
            hrefs.append(href)
    return hrefs


async def has_next_page(page):
    next_link = page.locator('a[rel="next"], a[href*="/page/"]')
    return await next_link.count() > 0


async def run_parser_job(job: Job):
    """Runs the full 2GIS grid scrape for job.queries, writing to job.raw_path.
    Mutates job.status/log/stats as it goes. Calls job.on_captcha(job) and
    pauses (awaiting resume) if a CAPTCHA page is detected."""
    os.makedirs(job.out_dir, exist_ok=True)
    job.status = "running"
    job.log(f"Старт парсинга ниши «{job.category}», запросов: {len(job.queries)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "База"
    ws.append(["№", "Название", "Адрес", "Телефон", "Сайт", "Тип сайта", "URL"])
    wb.save(job.raw_path)

    saved_ids = set()
    lock = asyncio.Lock()
    semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)
    grid_points = generate_grid(LAT_MIN, LAT_MAX, LON_MIN, LON_MAX, GRID_STEPS)

    display = os.environ.get("DISPLAY", ":99")
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            env={**os.environ, "DISPLAY": display},
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="ru-RU", timezone_id="Europe/Moscow", viewport={"width": 1280, "height": 720},
        )
        main_page = await context.new_page()

        try:
            for q_idx, item in enumerate(job.queries, 1):
                search_query = item["query"]
                keywords = item.get("keywords") or [w for w in re.split(r'\s+', search_query) if len(w) > 2]
                job.log(f"Запрос [{q_idx}/{len(job.queries)}]: '{search_query}'")
                encoded_query = urllib.parse.quote(search_query)

                for cell_idx, (lat, lon, sector_name) in enumerate(grid_points, 1):
                    page_num = 1
                    duplicate_pages = 0
                    while True:
                        search_url = f"https://2gis.ru/moscow/search/{encoded_query}/page/{page_num}?m={lon}%2C{lat}%2F{ZOOM}"
                        try:
                            nav_response = await main_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                            await main_page.wait_for_timeout(1200)

                            blocked_by_status = nav_response is not None and nav_response.status in CAPTCHA_STATUS_CODES
                            if blocked_by_status or await page_looks_like_captcha(main_page):
                                job.status = "captcha"
                                job.log("⚠️ Обнаружена капча/блокировка 2ГИС — жду ручного решения (до 8 минут)")
                                if job.on_captcha:
                                    await job.on_captcha(job)
                                cleared = False
                                for _ in range(32):  # 32 * 15s = 8 min
                                    await asyncio.sleep(15)
                                    reload_response = await main_page.reload(wait_until="domcontentloaded")
                                    still_blocked_by_status = reload_response is not None and reload_response.status in CAPTCHA_STATUS_CODES
                                    if not still_blocked_by_status and not await page_looks_like_captcha(main_page):
                                        cleared = True
                                        break
                                if not cleared:
                                    job.status = "error"
                                    job.log("🛑 Капча не решена вовремя, сектор пропущен")
                                    break
                                job.status = "running"
                                job.log("✅ Капча решена, продолжаю")

                            for _ in range(6):
                                current_count = await main_page.locator('a[href*="/firm/"]').count()
                                if current_count >= 24:
                                    break
                                await main_page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                                await main_page.wait_for_timeout(400)
                        except Exception as e:
                            job.log(f"Ошибка загрузки страницы: {e}")
                            break

                        hrefs = await collect_valid_hrefs(main_page)
                        if len(hrefs) == 0:
                            break

                        tasks = []
                        dup_count = 0
                        for href in hrefs:
                            firm_id = get_firm_id(href)
                            if not firm_id:
                                continue
                            if firm_id in saved_ids:
                                dup_count += 1
                                continue
                            clean = href.split('?')[0]
                            full_url = clean if clean.startswith('http') else f"https://2gis.ru{clean}"
                            saved_ids.add(firm_id)
                            tasks.append(process_firm(context, full_url, ws, wb, lock, semaphore, job, keywords))

                        async with lock:
                            job.stats["duplicates"] += dup_count

                        if tasks:
                            await asyncio.gather(*tasks, return_exceptions=True)

                        if not await has_next_page(main_page):
                            break
                        if dup_count == len(hrefs) and len(hrefs) > 0:
                            duplicate_pages += 1
                        else:
                            duplicate_pages = 0
                        if duplicate_pages >= 2:
                            break
                        page_num += 1
                        if page_num > 15:
                            break

                    if cell_idx % 10 == 0:
                        wb.save(job.raw_path)

                job.log(f"Готово: '{search_query}' — всего собрано {job.stats['collected']}")
        finally:
            wb.save(job.raw_path)
            await browser.close()

    job.status = "done"
    job.log(f"🏁 Парсинг завершён. Собрано: {job.stats['collected']}, дублей: {job.stats['duplicates']}, отфильтровано: {job.stats['filtered_out']}, ошибок: {job.stats['errors']}")


def get_domain(url):
    if not url:
        return None
    url_str = str(url).strip().lower()
    if url_str in ["нет сайта", "не указан"]:
        return None
    if not url_str.startswith('http'):
        url_str = 'http://' + url_str
    try:
        domain = urllib.parse.urlparse(url_str).netloc
        return domain[4:] if domain.startswith('www.') else domain
    except Exception:
        return None


def dedupe_franchises(raw_path, dedup_path):
    """Pure function version of clear_db.py's clean_franchises() - drops rows
    whose site domain repeats 2+ times (franchise chains), keeps the rest."""
    wb = openpyxl.load_workbook(raw_path)
    ws = wb.active

    domain_counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 5:
            domain = get_domain(row[4])
            if domain:
                domain_counts[domain] = domain_counts.get(domain, 0) + 1

    franchise_domains = {d for d, c in domain_counts.items() if c >= 2}

    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= 5 and get_domain(row[4]) in franchise_domains:
            rows_to_delete.append(i)
    for i in reversed(rows_to_delete):
        ws.delete_rows(i)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        row[0].value = i

    wb.save(dedup_path)
    return {"deleted_rows": len(rows_to_delete), "franchise_domains": len(franchise_domains)}


# --- LIVE RENDER (used by the hub's url-checker "live" scan mode) ---
#
# One-off job, unrelated to the 2GIS niche-scraping pipeline above: navigate
# to an arbitrary URL with the same headed-Chromium-under-Xvfb setup, wait for
# it to settle, best-effort decline a cookie-consent banner, and hand back the
# fully hydrated post-JS HTML. Reuses the same browser launch args/UA as
# run_parser_job for consistency (same anti-automation-detection posture).

RENDER_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


async def find_and_click_cookie_decline(page):
    """Best-effort, in-page search for a visible control whose text confidently
    reads as a cookie-consent decline/reject action, and click it. Only ever
    clicks when it finds a text match against an explicit decline-ish phrase
    list - never guesses at an "accept"/"ok" button, and if a banner is seen
    but no confident decline control is found, just reports that instead of
    clicking something that might actually be an accept."""
    try:
        return await page.evaluate("""() => {
            const declineMarkers = [
                "отклонить всё", "отклонить все", "отклонить", "запретить все", "запретить",
                "только необходимые", "только необходимые файлы", "не принимать",
                "reject all", "decline all", "reject", "decline",
                "necessary only", "necessary cookies only",
            ];
            const bannerMarkers = [
                "cookie", "куки", "cookie-файл", "файлов cookie", "персональных данных",
            ];

            const nodes = Array.from(document.querySelectorAll(
                'button, a, [role="button"], input[type="button"], input[type="submit"]'
            ));
            const visible = nodes.filter(el => {
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            });

            let bannerSeen = false;
            let target = null;
            let targetText = '';
            for (const el of visible) {
                const raw = el.innerText || el.value || el.textContent || '';
                const text = raw.trim().toLowerCase();
                if (!text || text.length > 60) continue;
                if (bannerMarkers.some(m => text.includes(m))) bannerSeen = true;
                if (!target && declineMarkers.some(m => text === m || text.includes(m))) {
                    target = el;
                    targetText = raw.trim().slice(0, 80);
                }
            }
            if (!target) {
                return {
                    detected: bannerSeen,
                    clicked: false,
                    note: bannerSeen
                        ? "cookie banner text seen but no confidently-labeled decline control found"
                        : "no cookie banner detected",
                };
            }
            try {
                target.click();
                return { detected: true, clicked: true, buttonText: targetText };
            } catch (e) {
                return { detected: true, clicked: false, buttonText: targetText, note: `click failed: ${e}` };
            }
        }""")
    except Exception as e:
        return {"detected": False, "clicked": False, "note": f"cookie banner inspection failed: {e}"}


async def render_live_page(url, timeout_ms=25000):
    """Launches a fresh headed Chromium (via Xvfb), navigates to `url`, waits
    for it to settle (network-idle, falling back to a fixed grace period for
    SPAs that keep background network activity alive forever), best-effort
    declines a cookie-consent banner, and returns the fully hydrated HTML plus
    what was found/done about the banner. Used for a single ad-hoc check, not
    queued through the niche-scraping job system."""
    display = os.environ.get("DISPLAY", ":99")
    result = {
        "ok": False,
        "url": url,
        "final_url": url,
        "status": None,
        "html": "",
        "headers": {},
        "cookie_banner": {"detected": False, "clicked": False},
    }

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            env={**os.environ, "DISPLAY": display},
            args=["--disable-blink-features=AutomationControlled"],
        )
        try:
            context = await browser.new_context(
                user_agent=RENDER_USER_AGENT,
                locale="ru-RU", timezone_id="Europe/Moscow", viewport={"width": 1366, "height": 900},
            )
            page = await context.new_page()
            nav_response = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)

            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass  # long-polling/analytics can keep the network busy forever - HTML is likely settled anyway

            await page.wait_for_timeout(800)  # let post-hydration banners/widgets finish mounting

            result["cookie_banner"] = await find_and_click_cookie_decline(page)
            if result["cookie_banner"].get("clicked"):
                await page.wait_for_timeout(600)  # let the banner's dismiss animation/DOM update settle

            result["html"] = await page.content()
            result["final_url"] = page.url
            if nav_response is not None:
                result["status"] = nav_response.status
                try:
                    result["headers"] = dict(nav_response.headers)
                except Exception:
                    result["headers"] = {}
            result["ok"] = True
        finally:
            await browser.close()

    return result
